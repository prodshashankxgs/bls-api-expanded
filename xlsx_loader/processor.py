#!/usr/bin/env python3
"""
BLS Excel Data Processor
========================

Processes CPI Excel files downloaded from BLS supplemental files.
Extracts seasonally adjusted and non-seasonally adjusted index numbers
for the most recent 2 months.
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any

import pandas as pd
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelDataProcessor:
    """
    Processes BLS CPI Excel files to extract index data
    """
    
    def __init__(self, data_sheet_dir: str = "data_sheet"):
        """
        Initialize processor
        
        Args:
            data_sheet_dir: Directory containing Excel files
        """
        self.data_sheet_dir = Path(data_sheet_dir)
        
        # Common CPI series patterns to look for
        self.cpi_series_patterns = {
            'cpi_u_sa': {
                'patterns': ['CUUR0000SA0', 'CPI-U.*seasonally adjusted', 'All items.*SA'],
                'description': 'CPI-U All Items Seasonally Adjusted',
                'seasonally_adjusted': True
            },
            'cpi_u_nsa': {
                'patterns': ['CUUR0000SA0', 'CPI-U.*not seasonally adjusted', 'All items.*NSA', 'All items(?!.*SA)'],
                'description': 'CPI-U All Items Not Seasonally Adjusted',
                'seasonally_adjusted': False
            }
        }
    
    def extract_cpi_data(self, excel_file: Path, ticker: str = "cpi", target_date: str = None) -> List[Dict]:
        """
        Extract CPI data from Excel file
        
        Args:
            excel_file: Path to Excel file
            ticker: Data ticker (e.g., 'cpi')
            target_date: Target date (e.g., '2025-06')
            
        Returns:
            List of data dictionaries compatible with existing data_loader format
        """
        try:
            logger.info(f"Processing Excel file: {excel_file}")
            
            # Load workbook
            workbook = load_workbook(excel_file, data_only=True)
            
            all_data = []
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")
                
                worksheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_data(worksheet, ticker, target_date)
                
                if sheet_data:
                    all_data.extend(sheet_data)
                    logger.info(f"Extracted {len(sheet_data)} data points from {sheet_name}")
            
            # Filter to most recent 2 months
            filtered_data = self._filter_recent_months(all_data, 2)
            
            logger.info(f"Total extracted data points: {len(filtered_data)}")
            return filtered_data
            
        except Exception as e:
            logger.error(f"Error processing Excel file {excel_file}: {e}")
            return []
    
    def _extract_sheet_data(self, worksheet, ticker: str, target_date: str) -> List[Dict]:
        """
        Extract data from a single worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            ticker: Data ticker
            target_date: Target date
            
        Returns:
            List of data dictionaries
        """
        try:
            # Convert worksheet to pandas DataFrame for easier processing
            data = []
            
            # Read all data from worksheet
            for row in worksheet.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                return []
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Try different extraction methods
            extracted_data = []
            
            # Method 1: Look for standard BLS table format
            method1_data = self._extract_standard_format(df, ticker)
            if method1_data:
                extracted_data.extend(method1_data)
            
            # Method 2: Look for time series format
            method2_data = self._extract_timeseries_format(df, ticker)
            if method2_data:
                extracted_data.extend(method2_data)
            
            # Method 3: Look for pivot table format
            method3_data = self._extract_pivot_format(df, ticker)
            if method3_data:
                extracted_data.extend(method3_data)
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"Error extracting sheet data: {e}")
            return []
    
    def _extract_standard_format(self, df: pd.DataFrame, ticker: str) -> List[Dict]:
        """
        Extract data from standard BLS table format
        
        Expected format:
        Year | Jan | Feb | Mar | ... | Dec
        2024 | 310.3 | 311.1 | ... | ...
        """
        try:
            data_points = []
            
            # Look for year column and month columns
            year_col = None
            month_cols = {}
            
            # Search for headers
            for idx, row in df.iterrows():
                if idx > 10:  # Don't search too far down
                    break
                
                row_str = ' '.join(str(cell).lower() for cell in row if pd.notna(cell))
                
                # Look for year identifier
                if 'year' in row_str or any(str(cell).isdigit() and len(str(cell)) == 4 for cell in row if pd.notna(cell)):
                    # This might be a header row
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).lower() if pd.notna(cell) else ''
                        
                        if 'year' in cell_str:
                            year_col = col_idx
                        elif any(month in cell_str for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                                               'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
                            month_name = cell_str[:3]
                            month_num = self._month_name_to_number(month_name)
                            if month_num:
                                month_cols[col_idx] = month_num
                    
                    if year_col is not None and month_cols:
                        # Process data rows starting from next row
                        for data_idx in range(idx + 1, len(df)):
                            data_row = df.iloc[data_idx]
                            
                            # Get year
                            year_val = data_row.iloc[year_col] if year_col < len(data_row) else None
                            if not year_val or not str(year_val).isdigit():
                                continue
                            
                            year = int(year_val)
                            
                            # Process each month
                            for col_idx, month in month_cols.items():
                                if col_idx < len(data_row):
                                    value = data_row.iloc[col_idx]
                                    if pd.notna(value) and self._is_numeric(value):
                                        data_points.append({
                                            'ticker': ticker,
                                            'year': year,
                                            'month': month,
                                            'value': float(value),
                                            'series_id': f'CPI_U_{ticker.upper()}',
                                            'category': 'Consumer Price Index (CPI-U)',
                                            'source': 'bls_excel',
                                            'seasonally_adjusted': True  # Default assumption
                                        })
                        break
            
            return data_points
            
        except Exception as e:
            logger.debug(f"Standard format extraction failed: {e}")
            return []
    
    def _extract_timeseries_format(self, df: pd.DataFrame, ticker: str) -> List[Dict]:
        """
        Extract data from time series format
        
        Expected format:
        Date | Value | Series
        2024-01 | 310.3 | CUUR0000SA0
        """
        try:
            data_points = []
            
            # Look for date and value columns
            date_col = None
            value_col = None
            series_col = None
            
            # Search first few rows for headers
            for idx in range(min(10, len(df))):
                row = df.iloc[idx]
                
                for col_idx, cell in enumerate(row):
                    cell_str = str(cell).lower() if pd.notna(cell) else ''
                    
                    if any(word in cell_str for word in ['date', 'period', 'time']):
                        date_col = col_idx
                    elif any(word in cell_str for word in ['value', 'index', 'cpi']):
                        value_col = col_idx
                    elif any(word in cell_str for word in ['series', 'id', 'code']):
                        series_col = col_idx
                
                if date_col is not None and value_col is not None:
                    # Process data rows
                    for data_idx in range(idx + 1, len(df)):
                        data_row = df.iloc[data_idx]
                        
                        if len(data_row) <= max(date_col, value_col):
                            continue
                        
                        date_val = data_row.iloc[date_col]
                        value_val = data_row.iloc[value_col]
                        
                        if pd.notna(date_val) and pd.notna(value_val) and self._is_numeric(value_val):
                            # Parse date
                            year, month = self._parse_date(date_val)
                            if year and month:
                                series_id = data_row.iloc[series_col] if series_col is not None and series_col < len(data_row) else ''
                                
                                data_points.append({
                                    'ticker': ticker,
                                    'year': year,
                                    'month': month,
                                    'value': float(value_val),
                                    'series_id': str(series_id) if pd.notna(series_id) else f'CPI_U_{ticker.upper()}',
                                    'category': 'Consumer Price Index (CPI-U)',
                                    'source': 'bls_excel',
                                    'seasonally_adjusted': self._determine_seasonal_adjustment(str(series_id))
                                })
                    break
            
            return data_points
            
        except Exception as e:
            logger.debug(f"Time series format extraction failed: {e}")
            return []
    
    def _extract_pivot_format(self, df: pd.DataFrame, ticker: str) -> List[Dict]:
        """
        Extract data from pivot table format where data might be arranged differently
        """
        try:
            data_points = []
            
            # Look for any numeric data with year/month patterns
            for idx, row in df.iterrows():
                for col_idx, cell in enumerate(row):
                    if pd.notna(cell) and self._is_numeric(cell):
                        # Look around this cell for year/month information
                        year, month = self._find_year_month_context(df, idx, col_idx)
                        
                        if year and month:
                            data_points.append({
                                'ticker': ticker,
                                'year': year,
                                'month': month,
                                'value': float(cell),
                                'series_id': f'CPI_U_{ticker.upper()}',
                                'category': 'Consumer Price Index (CPI-U)',
                                'source': 'bls_excel',
                                'seasonally_adjusted': True  # Default assumption
                            })
            
            return data_points
            
        except Exception as e:
            logger.debug(f"Pivot format extraction failed: {e}")
            return []
    
    def _find_year_month_context(self, df: pd.DataFrame, row_idx: int, col_idx: int) -> Tuple[Optional[int], Optional[int]]:
        """
        Find year and month context around a specific cell
        """
        try:
            year = None
            month = None
            
            # Search in the same row for year
            row = df.iloc[row_idx]
            for cell in row:
                if pd.notna(cell) and str(cell).isdigit() and len(str(cell)) == 4:
                    potential_year = int(cell)
                    if 2020 <= potential_year <= 2030:  # Reasonable year range
                        year = potential_year
                        break
            
            # Search in the same column for year if not found
            if year is None:
                col = df.iloc[:, col_idx] if col_idx < df.shape[1] else pd.Series()
                for cell in col:
                    if pd.notna(cell) and str(cell).isdigit() and len(str(cell)) == 4:
                        potential_year = int(cell)
                        if 2020 <= potential_year <= 2030:
                            year = potential_year
                            break
            
            # Search for month in headers (first few rows)
            if col_idx < df.shape[1]:
                for header_idx in range(min(5, len(df))):
                    header_cell = df.iloc[header_idx, col_idx]
                    if pd.notna(header_cell):
                        month = self._extract_month_from_text(str(header_cell))
                        if month:
                            break
            
            return year, month
            
        except Exception as e:
            logger.debug(f"Error finding year/month context: {e}")
            return None, None
    
    def _extract_month_from_text(self, text: str) -> Optional[int]:
        """Extract month number from text"""
        try:
            text = text.lower()
            month_map = {
                'jan': 1, 'january': 1,
                'feb': 2, 'february': 2,
                'mar': 3, 'march': 3,
                'apr': 4, 'april': 4,
                'may': 5,
                'jun': 6, 'june': 6,
                'jul': 7, 'july': 7,
                'aug': 8, 'august': 8,
                'sep': 9, 'september': 9, 'sept': 9,
                'oct': 10, 'october': 10,
                'nov': 11, 'november': 11,
                'dec': 12, 'december': 12
            }
            
            for month_name, month_num in month_map.items():
                if month_name in text:
                    return month_num
            
            return None
            
        except:
            return None
    
    def _month_name_to_number(self, month_name: str) -> Optional[int]:
        """Convert month name to number"""
        return self._extract_month_from_text(month_name)
    
    def _parse_date(self, date_val: Any) -> Tuple[Optional[int], Optional[int]]:
        """Parse date value to year and month"""
        try:
            date_str = str(date_val)
            
            # Try different date formats
            if '-' in date_str:
                parts = date_str.split('-')
                if len(parts) >= 2:
                    year = int(parts[0]) if parts[0].isdigit() else None
                    month = int(parts[1]) if parts[1].isdigit() else None
                    return year, month
            
            # Try pandas date parsing
            try:
                date_obj = pd.to_datetime(date_val)
                return date_obj.year, date_obj.month
            except:
                pass
            
            return None, None
            
        except:
            return None, None
    
    def _is_numeric(self, value: Any) -> bool:
        """Check if value is numeric"""
        try:
            float(value)
            return True
        except:
            return False
    
    def _determine_seasonal_adjustment(self, series_id: str) -> bool:
        """Determine if series is seasonally adjusted based on series ID"""
        series_id = series_id.upper()
        
        # BLS convention: SA = Seasonally Adjusted, NSA = Not Seasonally Adjusted
        if 'SA' in series_id and 'NSA' not in series_id:
            return True
        elif 'NSA' in series_id:
            return False
        
        # Default assumption for CPI-U
        return True
    
    def _filter_recent_months(self, data: List[Dict], num_months: int = 2) -> List[Dict]:
        """
        Filter data to most recent N months
        
        Args:
            data: List of data dictionaries
            num_months: Number of recent months to keep
            
        Returns:
            Filtered data for recent months
        """
        try:
            if not data:
                return []
            
            # Sort by year and month (most recent first)
            sorted_data = sorted(data, key=lambda x: (x.get('year', 0), x.get('month', 0)), reverse=True)
            
            # Get unique year-month combinations
            seen_periods = set()
            filtered_data = []
            
            for item in sorted_data:
                year = item.get('year')
                month = item.get('month')
                
                if year and month:
                    period = (year, month)
                    if period not in seen_periods and len(seen_periods) < num_months:
                        seen_periods.add(period)
                    
                    if period in seen_periods:
                        filtered_data.append(item)
            
            logger.info(f"Filtered to {len(filtered_data)} data points for {len(seen_periods)} recent periods")
            
            return filtered_data
            
        except Exception as e:
            logger.error(f"Error filtering recent months: {e}")
            return data  # Return original data if filtering fails
    
    def get_file_info(self, excel_file: Path) -> Dict:
        """
        Get information about an Excel file
        
        Args:
            excel_file: Path to Excel file
            
        Returns:
            Dictionary with file information
        """
        try:
            workbook = load_workbook(excel_file, data_only=True)
            
            return {
                'filename': excel_file.name,
                'sheets': workbook.sheetnames,
                'sheet_count': len(workbook.sheetnames),
                'file_size': excel_file.stat().st_size,
                'modified_date': datetime.fromtimestamp(excel_file.stat().st_mtime)
            }
            
        except Exception as e:
            logger.error(f"Error getting file info for {excel_file}: {e}")
            return {}


if __name__ == "__main__":
    # Test the processor
    logging.basicConfig(level=logging.INFO)
    
    processor = ExcelDataProcessor()
    
    # Test with any Excel file in data_sheet folder
    data_sheet_path = Path("data_sheet")
    if data_sheet_path.exists():
        excel_files = list(data_sheet_path.glob("*.xlsx"))
        if excel_files:
            test_file = excel_files[0]
            print(f"Testing with file: {test_file}")
            data = processor.extract_cpi_data(test_file)
            print(f"Extracted {len(data)} data points")
            
            if data:
                print("Sample data:")
                for item in data[:5]:
                    print(f"  {item}")
        else:
            print("No Excel files found in data_sheet folder")
    else:
        print("data_sheet folder not found")